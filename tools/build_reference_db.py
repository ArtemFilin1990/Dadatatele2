from __future__ import annotations

import sqlite3
import zipfile
from pathlib import Path

TABLES = {
    "okved2": ["code", "title"],
    "okopf": ["code", "title"],
    "okfs": ["code", "title"],
    "okpd": ["code", "title"],
    "okpd2": ["code", "title"],
    "account_codes": ["code", "title"],
    "statuses": ["code", "title"],
}

ZIP_MAP = {
    "okved_2.sql.zip": "okved2",
    "okopf.sql.zip": "okopf",
    "okfs.sql.zip": "okfs",
    "okpd.sql.zip": "okpd",
    "okpd_2.sql.zip": "okpd2",
    "account_codes.sql.zip": "account_codes",
}


def create_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS okved2 (code TEXT PRIMARY KEY, title TEXT);
        CREATE TABLE IF NOT EXISTS okopf (code TEXT PRIMARY KEY, title TEXT);
        CREATE TABLE IF NOT EXISTS okfs (code TEXT PRIMARY KEY, title TEXT);
        CREATE TABLE IF NOT EXISTS okpd (code TEXT PRIMARY KEY, title TEXT);
        CREATE TABLE IF NOT EXISTS okpd2 (code TEXT PRIMARY KEY, title TEXT);
        CREATE TABLE IF NOT EXISTS account_codes (code TEXT PRIMARY KEY, title TEXT);
        CREATE TABLE IF NOT EXISTS statuses (code TEXT PRIMARY KEY, title TEXT);
        """
    )


def import_sql_zip(conn: sqlite3.Connection, zip_path: Path, table: str) -> None:
    with zipfile.ZipFile(zip_path) as zf:
        for name in zf.namelist():
            if name.lower().endswith(".sql"):
                sql_text = zf.read(name).decode("utf-8", errors="ignore")
                try:
                    conn.executescript(sql_text)
                except sqlite3.DatabaseError:
                    # fallback: simple parser for lines like code;title
                    for line in sql_text.splitlines():
                        if ";" not in line:
                            continue
                        left, right = line.split(";", 1)
                        code = left.strip().strip("'\"")
                        title = right.strip().strip("'\"")
                        if code and title:
                            conn.execute(
                                f"INSERT OR IGNORE INTO {table}(code,title) VALUES(?,?)",
                                (code, title),
                            )


def import_statuses_xlsx(conn: sqlite3.Connection, zip_path: Path) -> None:
    try:
        from openpyxl import load_workbook
    except Exception:
        return

    with zipfile.ZipFile(zip_path) as zf:
        for name in zf.namelist():
            if name.lower().endswith(".xlsx"):
                temp = Path("/tmp") / Path(name).name
                temp.write_bytes(zf.read(name))
                wb = load_workbook(temp)
                for ws in wb.worksheets:
                    for row in ws.iter_rows(min_row=1, values_only=True):
                        if not row or row[0] is None or row[1] is None:
                            continue
                        code = str(row[0]).strip()
                        title = str(row[1]).strip()
                        conn.execute("INSERT OR IGNORE INTO statuses(code,title) VALUES(?,?)", (code, title))


def build() -> None:
    db_dir = Path("db")
    db_dir.mkdir(parents=True, exist_ok=True)
    db_path = db_dir / "reference_data.sqlite"
    cache_path = db_dir / "cache.sqlite"

    assets_dir = Path("assets/reference_sources")
    with sqlite3.connect(db_path) as conn:
        create_schema(conn)
        if assets_dir.exists():
            for zip_name, table in ZIP_MAP.items():
                zp = assets_dir / zip_name
                if zp.exists():
                    import_sql_zip(conn, zp, table)
            status_zip = assets_dir / "statuses.xlsx.zip"
            if status_zip.exists():
                import_statuses_xlsx(conn, status_zip)
        conn.commit()

    sqlite3.connect(cache_path).close()
    print(f"built: {db_path} and {cache_path}")


if __name__ == "__main__":
    build()
